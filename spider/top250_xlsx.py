import re
import requests
from bs4 import BeautifulSoup
import openpyxl
import datetime
from tools.random_sleep import random_sleep_time

'''
######生成随机睡眠数函数random_sleep_time
'''

'''
######从URL中获取主要的内容
'''
def get_url(url,header):
    try:
        random_sleep_time()
        con = requests.get(url,headers=header)
        if con.status_code is 200:
            return con.text
        else:
            return "error"
    except:
        return "error"

'''
######解析网页body，并提取需求的内容
'''
def parse_li(li,moive):
    moive['title'] = li.find("span",attrs={'class':'title'}).text
    moive['rating_num'] = li.find("span",attrs={'class':'rating_num'}).text
    moive['href'] = li.find("a",attrs={'class',''})['href']
    try:
        moive['quto'] = (li.find("span",attrs={'class','inq'}).text)
    except:
        moive['quto'] = "【没有quto】"

    print("完成:#######"+moive['title']+"的xlsx格式保存")
    return moive
'''
######主函数，书要是将top250的电影按排名存入到top250.xlsx的文件中
'''
def to_xlsx():
    url = "https://movie.douban.com/top250?start="    
    agent = "Mozilla/5.0 (Windows NT 6.1; rv:59.0) Gecko/20100101 Firefox/59.0"
    header = {'User-Agent':agent}    
    wb = openpyxl.Workbook()
    ws =wb.active
    ws['A1']="TITLE"
    ws['B1']="评分"
    ws['C1']="href"
    ws['D1']="quto"
    nums = 2
    moive={}
    for i in range(10):
        new_url = url+str(i*25)+"&filter"
        con = get_url(new_url,header)
        if con is 'error':
            continue
        con = BeautifulSoup(con,"lxml")
        con_l = con.ol.find_all("li")
        for each_li in con_l:
            moive=parse_li(each_li,moive)
            ws['A'+str(nums)]  = moive['title']
            ws['B'+str(nums)] = moive['rating_num']
            ws['C'+str(nums)] = moive['href']
            ws['D'+str(nums)] = moive['quto']
            nums=nums+1

    wb.save("doubanTOP250.xlsx")
    
        
'''
#######运行代码
'''
    

